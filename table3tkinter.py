import tkinter as tk
import openpyxl
from openpyxl.workbook import Workbook
from openpyxl import load_workbook

# Open the Excel file and get the active worksheet
workbook = openpyxl.load_workbook('zayifkiloalma.xlsx')
sheet = workbook.active

# Create a tkinter root window
root = tk.Tk()

# Iterate through the rows of the worksheet and print the values
for row in sheet.rows:
    print(row[0].value, row[1].value, row[2].value)

# Start the tkinter main loop
root.mainloop()





#CALISAN
    if vki < 18.5:
        messagebox.showinfo('vucud_kitle_indeksi_hesaplama', f'vki = {vki} Kilonuz zayif.')
        root = tk.Tk()
        root.title('ZayifProgram')
        file ="zayifkiloalma.xlsx"
        wb = load_workbook(file,data_only=TRUE)
        ws= wb.active

        r=0
        for row in ws:
            c = 0 
            for cell in row:
                tk.Label(root,text=cell.value).grid(row=r,column=c)
                c+=1
            r+=1
        root.mainloop()        